# Importing necessary modules after ensuring they are available
import json
import pandas as pd
import warnings
from colorama import Fore, Style
from openpyxl import Workbook

# Setting a filter to ignore warnings during execution
warnings.simplefilter("ignore")

data_file_name  = "data.json"
try:
    with open(data_file_name) as file:
        file_data_file_name_data = json.load(file)
        permitted_list = json.dumps(file_data_file_name_data['equipments'])
except Exception as e:
    raise Exception(f'\nCrie o arquivo "data.json".\n\nOcorreu um erro: {e}')


def map_file(map_type, calculated=False):
    """
    Generates the filename based on the map type and calculation status.

    Args:
        map_type (str): Type of map, either 'infraction' or 'vehicles'.
        calculated (bool): Specifies whether the map is calculated or not.

    Returns:
        str: Filename.
    """
    suffix_calculate = " - Calculated"
    suffix_file = ".xlsx"
    if map_type == 'infraction':
        name = 'Mapa de infrações'
    elif map_type == 'vehicles':
        name = 'Mapa de veículos'

    if calculated:
        file_name = name + suffix_calculate + suffix_file
        return file_name
    if not calculated:
        file_name = name + suffix_file
        return file_name


def process_file(map_type):
    """
    Processes the Excel file and extracts data based on the map type.

    Args:
        map_type (str): Type of map, either 'infraction' or 'vehicles'.

    Returns:
        list: Processed data.
    """
    if map_type == 'infraction':
        col_idx = 11
    elif map_type == 'vehicles':
        col_idx = 13

    with pd.ExcelFile(map_file(map_type)) as xls:
        print(f' >| {Fore.LIGHTWHITE_EX}Processing {Fore.LIGHTRED_EX}"{map_file(map_type)}"{Style.RESET_ALL}')
        sheets = xls.sheet_names
        data = []
        id = 0
        id_s = 0
        for sheet_name in sheets:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=None, skiprows=8, nrows=32)
            equipment = df.at[0, col_idx]
            lane = df.at[1, col_idx]
            id_s += 1
            id_s_p = "{:02d}".format(id_s)
            last_detected = ""
            if map_type == 'vehicles':
                equipment = equipment[:9]

            if equipment in permitted_list:
                id += 1
                id_p = "{:02d}".format(id)
                last_detected = f' >| {Fore.GREEN}Last Sheet Processed: {id_p} - {equipment} - {lane}{Style.RESET_ALL} |<'
                entry = [equipment, lane]
                for j in range(2, df.shape[1]):
                    value = df.at[31, j]
                    if not pd.isnull(value):
                        entry.append(value)
                data.append(entry)
            print(f'\r >| {Fore.CYAN}Last Sheet Found: {id_s_p} - {equipment} - {lane}{Style.RESET_ALL} |<{last_detected}{Style.RESET_ALL}', end=' ', flush=True)

            continue
        print()

        columns = list(zip(*data))
        for i, column in enumerate(columns):
            if all(value == 0 for value in column[2:-1]):  # Checks if all values except first and last are 0
                columns[i] = tuple("" if value == 0 else value for value in column)
        result = list(zip(*columns))
        return result


def save_file(map_type):
    """
    Saves processed data to a new Excel file.

    Args:
        map_type (str): Type of map, either 'infraction' or 'vehicles'.
    """
    map_type_boolean = True
    map_file_name = map_file(map_type, map_type_boolean)

    data = process_file(map_type)
    print(f'\r >|{Fore.LIGHTWHITE_EX}Writing File {Fore.LIGHTRED_EX}"{map_file_name}"{Style.RESET_ALL}', end=' ', flush=True)

    wb = Workbook()
    ws = wb.active

    # Defining the header row for the new Excel file
    header = ["Equipment", "Lane"]
    for i in range(2, len(data[0]) - 1):
        header.append(i - 1)
    header.append("Total")

    # Writing the header row to the worksheet
    for i, val in enumerate(header, start=1):
        ws.cell(row=1, column=i, value=val)

    # Writing the extracted data to the worksheet
    for row_num, row_data in enumerate(data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

    print(f'\r >| {Fore.LIGHTWHITE_EX}Writed File {Fore.LIGHTRED_EX}"{map_file_name}"{Style.RESET_ALL}', end=' ', flush=True)
    # Saving the new Excel workbook with processed data
    wb.save(map_file_name)

save_file('infraction')
print('\n')
save_file('vehicles')
print('\n')

# answer = input("\nPress 'ENTER' to exit...")
# if answer:
#     sys.exit()
